"""
Excel Exporter — plain output, no colors, no author metadata.
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── styles ────────────────────────────────────────────────────────────────────
BOLD   = Font(name="Arial", bold=True,  size=10)
NORMAL = Font(name="Arial", bold=False, size=10)
ITALIC = Font(name="Arial", italic=True, size=9, color="666666")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")

def _border():
    s = Side(style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)

def _header(cell, text):
    cell.value     = text
    cell.font      = BOLD
    cell.alignment = CENTER
    cell.border    = _border()

def _cell(cell, value, align=CENTER):
    cell.value     = value
    cell.font      = NORMAL
    cell.alignment = align
    cell.border    = _border()

def _no_meta(wb):
    wb.properties.creator        = ""
    wb.properties.lastModifiedBy = ""
    wb.properties.company        = ""
    wb.properties.description    = ""
    wb.properties.subject        = ""
    wb.properties.title          = ""
    wb.properties.keywords       = ""

# ═══════════════════════════════════════════════════════════════════════════════
# COLUMN DEFINITIONS
# ── (Excel header label, data field name, column width) ──────────────────────
# تعديل المسميات: غيّر النص في العمود الأول فقط
# ═══════════════════════════════════════════════════════════════════════════════

SUMMARY_COLS = [
    ("رقم الموظف",          "رقم الموظف",          12),
    ("اسم الموظف",          "اسم الموظف",           20),
    ("المديرية",             "المديرية",              22),
    ("أيام الدوام الفعلية",         "أيام_عمل_متوقعة",       13),
    ("أيام حضور",           "أيام_حضور",              12),
    ("غياب غير مبرر",       "غياب_غير_مبرر",          16),
    ("غياب مبرر",           "غياب_مبرر",              14),
    ("تأخير الدخول (HH:MM)",       "تأخير_hhmm",             15),
    ("خروج مبكر (HH:MM)",        "مبكر_hhmm",              15),
    ("إضافي صافي (HH:MM)",  "اضافي_صافي_hhmm",        18),
    ("معدل الحضور %",       "معدل_الحضور",             15),
]

DETAIL_COLS = [
    ("رقم الموظف",      "رقم الموظف",        12),
    ("اسم الموظف",      "اسم الموظف",         18),
    ("التاريخ",         "التاريخ",             14),
    ("الشهر",           "الشهر",              10),
    ("أول دخول",        "أول دخول",           10),
    ("آخر خروج",        "آخر خروج",           10),
    ("الحالة الفعلية",  "الحالة_الفعلية",     18),
    ("تأخير خام (د)",   "تأخير_خام",          14),
    ("تأخير محسوب (د)", "تأخير_مقرب",         15),
    ("مبرر تأخير",      "تأخير_مبرر",         12),
    ("خروج مبكر خام (د)",    "مبكر_خام",           13),
    ("خروج مبكر محسوب (د)",  "مبكر_مقرب",          14),
    ("إضافي صافي (د)",  "اضافي_صافي",         14),
    ("مدة العمل (د)",   "مدة_العمل_دقيقة",    14),
    ("ملاحظة",          "ملاحظة",             24),
]

# ── Summary sheet ─────────────────────────────────────────────────────────────

def _write_summary(wb, summary):
    ws = wb.create_sheet("ملخص الموظفين")
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    for c, (label, _, width) in enumerate(SUMMARY_COLS, 1):
        ws.column_dimensions[get_column_letter(c)].width = width
        _header(ws.cell(1, c), label)

    for r, (_, row) in enumerate(summary.iterrows(), 2):
        for c, (_, field, _) in enumerate(SUMMARY_COLS, 1):
            val = row.get(field, "")
            if isinstance(val, float) and str(val) == "nan":
                val = ""
            cell = ws.cell(r, c, val)
            cell.font      = NORMAL
            cell.alignment = CENTER
            cell.border    = _border()
            if field == "معدل_الحضور":
                cell.number_format = "0.0%"

    ws.auto_filter.ref = f"A1:{get_column_letter(len(SUMMARY_COLS))}1"

# ── Detail sheet ──────────────────────────────────────────────────────────────

def _write_detail(wb, detail):
    ws = wb.create_sheet("التفصيل اليومي")
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    for c, (label, _, width) in enumerate(DETAIL_COLS, 1):
        ws.column_dimensions[get_column_letter(c)].width = width
        _header(ws.cell(1, c), label)

    for r, (_, row) in enumerate(detail.iterrows(), 2):
        for c, (_, field, _) in enumerate(DETAIL_COLS, 1):
            val = row.get(field, "")
            if isinstance(val, float) and str(val) == "nan":
                val = ""
            if isinstance(val, bool):
                val = "نعم" if val else ""
            cell = ws.cell(r, c, val)
            cell.font      = NORMAL
            cell.alignment = CENTER
            cell.border    = _border()

    ws.auto_filter.ref = f"A1:{get_column_letter(len(DETAIL_COLS))}1"

# ── Stats sheet ───────────────────────────────────────────────────────────────

def _write_stats(wb, detail, summary):
    ws = wb.create_sheet("إحصاءات")
    ws.sheet_view.rightToLeft = True
    working = detail[~detail["عطلة_رسمية"]]

    by_month = working.groupby("الشهر").agg(
        حضور          =("الحالة_الفعلية", lambda x: x.isin({"حاضر","خروج غير مسجل"}).sum()),
        غياب_مبرر     =("غياب_مبرر", "sum"),
        غياب_غير_مبرر =("الحالة_الفعلية", lambda x: (x=="غياب غير مبرر").sum()),
        اضافي_دقيقة   =("اضافي_صافي", "sum"),
        تأخير_دقيقة   =("تأخير_مقرب", "sum"),
    ).reset_index()

    for c, h in enumerate(["الشهر","حضور","غياب مبرر","غياب غير مبرر","إضافي (د)","تأخير (د)"], 1):
        ws.column_dimensions[get_column_letter(c)].width = 18
        _header(ws.cell(1, c), h)
    for r, (_, row) in enumerate(by_month.iterrows(), 2):
        for c, f in enumerate(["الشهر","حضور","غياب_مبرر","غياب_غير_مبرر","اضافي_دقيقة","تأخير_دقيقة"], 1):
            _cell(ws.cell(r, c), row.get(f, ""))

    start = len(by_month) + 4
    by_dept = working.groupby("المديرية").agg(
        عدد_الموظفين     =("رقم الموظف", "nunique"),
        غياب_غير_مبرر    =("الحالة_الفعلية", lambda x: (x=="غياب غير مبرر").sum()),
        اضافي_صافي_دقيقة =("اضافي_صافي", "sum"),
    ).reset_index()

    for c, h in enumerate(["المديرية","عدد الموظفين","غياب غير مبرر","إضافي صافي (د)"], 1):
        _header(ws.cell(start, c), h)
    for r, (_, row) in enumerate(by_dept.iterrows(), start + 1):
        for c, f in enumerate(["المديرية","عدد_الموظفين","غياب_غير_مبرر","اضافي_صافي_دقيقة"], 1):
            _cell(ws.cell(r, c), row.get(f, ""))

# ── Config sheet ──────────────────────────────────────────────────────────────

def _write_config(wb, cfg):
    ws = wb.create_sheet("الإعدادات المستخدمة")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 22
    _header(ws.cell(1, 1), "الإعداد")
    _header(ws.cell(1, 2), "القيمة")

    labels = {
        "shift_start":                       "وقت بداية الدوام",
        "shift_end":                         "وقت نهاية الدوام",
        "grace_minutes":                     "دقائق السماح",
        "tardiness_base":                    "احتساب التأخير من",
        "tardiness_rounding":                "طريقة تقريب التأخير",
        "tardiness_round_up_to":             "تقريب التأخير لأعلى (دقيقة)",
        "early_departure_tolerance_minutes": "تسامح المغادرة المبكرة (دقيقة)",
        "early_departure_round_up_to":       "تقريب المبكر لأعلى (دقيقة)",
        "overtime_threshold_minutes":        "حد الإضافي (دقيقة)",
        "overtime_round_down_to":            "تقريب الإضافي لأسفل (دقيقة)",
        "deduct_tardiness_from_overtime":    "طرح التأخير من الإضافي",
        "missing_checkin_action":            "حاضر بدون تسجيل دخول",
        "missing_checkout_action":           "حاضر بدون تسجيل خروج",
    }
    for i, (key, label) in enumerate(labels.items(), 2):
        _cell(ws.cell(i, 1), label, align=LEFT)
        _cell(ws.cell(i, 2), str(cfg.get(key, "")))

# ── Template ──────────────────────────────────────────────────────────────────

def build_template() -> bytes:
    """
    Exceptions template — single sheet.
    Date split into 3 columns (السنة / الشهر / اليوم) matching source YYYY/MM/DD order.
    3 justification types only. One example row clearly marked.
    """
    wb = Workbook()
    wb.remove(wb.active)
    _no_meta(wb)

    ws = wb.create_sheet("المبررات الفردية")
    ws.sheet_view.rightToLeft = True

    # column widths
    ws.column_dimensions["A"].width = 14   # رقم الموظف
    ws.column_dimensions["B"].width = 10   # السنة
    ws.column_dimensions["C"].width = 10   # الشهر
    ws.column_dimensions["D"].width = 10   # اليوم
    ws.column_dimensions["E"].width = 22   # النوع
    ws.column_dimensions["F"].width = 28   # ملاحظة

    # headers
    _header(ws.cell(1, 1), "رقم الموظف")
    _header(ws.cell(1, 2), "السنة")
    _header(ws.cell(1, 3), "الشهر (1-12)")
    _header(ws.cell(1, 4), "اليوم (1-31)")
    _header(ws.cell(1, 5), "النوع")
    _header(ws.cell(1, 6), "ملاحظة (اختياري)")

    # ── Data Validation ───────────────────────────────────────────────────
    # السنة
    dv_year = DataValidation(
        type="whole", operator="between",
        formula1="2020", formula2="2099",
        showErrorMessage=True,
        errorTitle="سنة غير صحيحة",
        error="أدخل سنة بين 2020 و 2099",
    )
    ws.add_data_validation(dv_year)
    dv_year.sqref = "B2:B1000"

    # الشهر
    dv_month = DataValidation(
        type="whole", operator="between",
        formula1="1", formula2="12",
        showErrorMessage=True,
        errorTitle="شهر غير صحيح",
        error="أدخل شهراً بين 1 و 12",
    )
    ws.add_data_validation(dv_month)
    dv_month.sqref = "C2:C1000"

    # اليوم
    dv_day = DataValidation(
        type="whole", operator="between",
        formula1="1", formula2="31",
        showErrorMessage=True,
        errorTitle="يوم غير صحيح",
        error="أدخل يوماً بين 1 و 31",
    )
    ws.add_data_validation(dv_day)
    dv_day.sqref = "D2:D1000"

    # النوع — قائمة منسدلة
    dv_type = DataValidation(
        type="list",
        formula1='"غياب مبرر,تأخير مبرر,خروج مبكر مبرر"',
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="نوع غير صحيح",
        error="اختر من القائمة: غياب مبرر / تأخير مبرر / خروج مبكر مبرر",
    )
    ws.add_data_validation(dv_type)
    dv_type.sqref = "E2:E1000"

    # ── صف مثال — مميز باللون ─────────────────────────────────────────────
    from openpyxl.styles import PatternFill as _PF
    ex_fill = _PF("solid", start_color="FFF2CC", fgColor="FFF2CC")
    ex_font = Font(name="Arial", size=9, italic=True, color="7F6000")

    example = [999, 2026, 4, 1, "غياب مبرر", "← مثال فقط — احذف هذا السطر قبل الرفع"]
    for c, val in enumerate(example, 1):
        cell = ws.cell(2, c, val)
        cell.fill      = ex_fill
        cell.font      = ex_font
        cell.alignment = CENTER if c < 6 else LEFT
        cell.border    = _border()

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Main report ───────────────────────────────────────────────────────────────

def build_excel_report(detail, summary, cfg) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    _no_meta(wb)

    _write_summary(wb, summary)
    _write_detail(wb, detail)
    _write_stats(wb, detail, summary)
    _write_config(wb, cfg)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
